'''
reader.py
'''
from openpyxl import load_workbook

# checks if cell has black bg, bold, and starts with 'Day'
def check_title(cell):
    return cell.value.startswith('Day') and cell.font.bold and cell.fill.start_color.index == 'FF000000'

def read(filename):
    wb = load_workbook(filename)
    ws = wb.active
    #print(wb.sheetnames)

    # iterate through sheets until 'Week 1' is found
    for sheet in wb.worksheets:
        if (sheet.title == 'Week 1'):
            print('found week 1:')
            sheet = wb.active
            find_title_cell(sheet)

# return data retrieved from singular sheet in the training program
# Example: Week 1 - Day 1 ... Day 2 ... Day 3 ...
def read_sheet(worksheet):
    

# return data retrieved from workout within a sheet
def read_workout(cell):
    # start at provided cell
    # iterate through rows and columns of singular workouts
    # record data to list/dict
    # return list/dict

